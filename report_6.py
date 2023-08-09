import os, sqlite3, openpyxl, datetime, sys
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QApplication
from PyQt5.QtWidgets import QGridLayout, QProgressBar, QWidget, QCheckBox
from PyQt5 import QtGui, QtCore

from data.design.report_6 import Ui_MainWindow as Ui_Report6

REPORT_TITLE = 'отчёт категорий'


class MyCheckBox(QCheckBox):
    def set_args(self, args):
        self.args = args


bold_big_style = Font(size="14", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")
little_style = Font(size="7")
pale_style = Font(size="9", color='00555555', italic=True)

all_month = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
             '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
             '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}


def get_period_name(date, year, type=2):
    if type == 1:
        return f'{all_month[date]}({date}) {year}', [date, year]
    if type == 2:
        return f'{(int(date) - 1) // 3 + 1}-й квартал {year} года', [(int(date) - 1) // 3 + 1, year]
    if type == 3:
        return str(year), [0, year]



def set_cell(sheet, x, y, value, style, alignment=1, fill=None, border=None):
    cell = sheet._get_cell(x, y)
    cell.value = value
    cell.font = style
    if alignment == 1:
        cell.alignment = Alignment(horizontal='left')
    elif alignment == 2:
        cell.alignment = Alignment(horizontal='center')
    elif alignment == 3:
        cell.alignment = Alignment(horizontal='right')
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border



def get_num_from_date(date):
    d, m, y = date.split('.')
    return int(m) * 31 + int(y) * 385 + int(d)


def get_date_calendar(calendar):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def get_datetime_date_from_calendar(calendar):
    calendar_date = calendar.selectedDate()
    datetime_date = datetime.date(day=calendar_date.day(), month=calendar_date.month(),
                         year=calendar_date.year())
    return datetime_date


class ReportMenu6(QMainWindow, Ui_Report6):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.button.clicked.connect(self.create_report)

        self.radio_month.clicked.connect(self.change_period)
        self.radio_quarter.clicked.connect(self.change_period)
        self.radio_year.clicked.connect(self.change_period)

        self.setWindowTitle('Настройка перед отчётом по категориям')

        self.change_period(type=2)

        font = QtGui.QFont()
        font.setPointSize(12)
        inquiry = f"""SELECT DISTINCT id, name FROM categories
                                    WHERE is_deleted = 0"""
        categories = self.cur.execute(inquiry).fetchall()
        for categori in categories:
            check_box = MyCheckBox(categori[1])
            check_box.set_args(categori[0])
            if categori[0] == 2 or categori[0] == 4:
                check_box.setChecked(True)
            check_box.clicked.connect(self.change_category)
            check_box.setFont(font)
            self.layout_categories.addWidget(check_box)
        self.chosen_category = [2, 4]

    def change_category(self, status):
        check_box = self.sender()
        if status:
            self.chosen_category.append(check_box.args)
        else:
            del self.chosen_category[self.chosen_category.index(check_box.args)]

    def get_text_rule_of_category(self):
        res = ''
        if len(self.chosen_category) == 0:
            return res
        elif len(self.chosen_category) == 1:
            return f'AND category = {self.chosen_category[0]}'
        res = f'AND (category = {self.chosen_category[0]}'
        for i in self.chosen_category[1:]:
            res += f' OR category = {i}'
        res += ')'
        return res

    def change_period(self, type=0):
        if type == 2:
            type = 2
        else:
            if self.radio_month.isChecked():
                type = 1
            elif self.radio_quarter.isChecked():
                type = 2
            else:
                type = 3

        inquiry = f"""SELECT DISTINCT date_of_operation FROM patients WHERE is_deleted = 0"""
        all_all_dates = self.cur.execute(inquiry).fetchall()
        self.all_dates = []
        for date in all_all_dates:
            date = '.'.join(date[0].split('.')[1::])
            num_date = date.split('.')
            name_date, num_date = get_period_name(*num_date, type)
            if (name_date, num_date) not in self.all_dates:
                self.all_dates.append((name_date, num_date))

        self.all_dates.sort(key=lambda x: int(x[1][0]) + int(x[1][1] * 12), reverse=True)
        self.selected_period.clear()
        for name in self.all_dates:
            self.selected_period.addItem(name[0])


    def create_report(self):
        if len(self.chosen_category) == 0:
            return
        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return
        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == REPORT_TITLE:
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        is_year = self.radio_year.isChecked()
        right_year = self.all_dates[self.selected_period.currentIndex()][1][1]
        if not is_year:
            if self.radio_quarter.isChecked():
                right_period = []
                quarter = self.all_dates[self.selected_period.currentIndex()][1][0]
                for i in range((quarter - 1) * 3 + 1, quarter * 3 + 1):
                    right_period.append(str(i).rjust(2, '0'))
            else:
                right_period = [self.all_dates[self.selected_period.currentIndex()][1][0].rjust(2, '0')]
        else:
            right_period = []

        sheet.title = REPORT_TITLE

        set_cell(sheet, 1, 1, 'Отчёт по категориям:', bold_big_style)
        if is_year:
            set_cell(sheet, 1, 3, f'За {self.selected_period.currentText()} год', bold_big_style)
        else:
            set_cell(sheet, 1, 3, f'За {self.selected_period.currentText()}', bold_big_style)
        if self.radio_all.isChecked():
            rule_patients = ''

            set_cell(sheet, 2, 1, '* данные о количестве процедур и усл.ед. беруться не по указанному промежутку времени, а по пациенту, у которого была дата запись в заданном промежутке (из-за этого данные могут несходиться с данными из других отчётов)', pale_style)
            inquiry = f"""SELECT DISTINCT * FROM patients
                   WHERE is_deleted = 0 {rule_patients} {self.get_text_rule_of_category()} AND date_of_operation LIKE '%{right_year}'"""
            all_patients = self.cur.execute(inquiry).fetchall()

        else:
            rule_patients = 'AND is_discharge = 1'

            set_cell(sheet, 2, 1, '* данные о количестве процедур и усл.ед. беруться не по указанному промежутку времени, а по пациенту, у которого была дата выписки в заданном промежутке (из-за этого данные могут несходиться с данными из других отчётов)', pale_style)
            inquiry = f"""SELECT DISTINCT * FROM patients
                               WHERE is_deleted = 0 {rule_patients} {self.get_text_rule_of_category()} AND date_of_discharge LIKE '%{right_year}'"""
            all_patients = self.cur.execute(inquiry).fetchall()

        num_str = 4
        all_patients.sort(key=lambda x: x[1])

        for category in self.chosen_category:
            global_sum_lessons = 0
            global_sum_price = 0
            count_right_patients = 0

            inquiry = f"""SELECT DISTINCT name FROM categories
            WHERE id = {category}"""
            name_category = self.cur.execute(inquiry).fetchall()[0][0]
            set_cell(sheet, num_str, 1, f'Категория "{name_category}"', bold_big_style)
            num_str += 1

            set_cell(sheet, num_str, 1, '№ истории', bold_style)
            set_cell(sheet, num_str, 2, 'ФИО', bold_style)
            set_cell(sheet, num_str, 3, 'кол.процедур', bold_style)
            set_cell(sheet, num_str, 4, 'кол.усл.ед.', bold_style)

            for patient in all_patients:
                month = patient[10].split('.')[1]
                if month not in right_period and not is_year or patient[4] != category:
                    continue
                count_right_patients += 1
                num_str += 1
                set_cell(sheet, num_str, 1, str(patient[3]), style, 2)
                set_cell(sheet, num_str, 2, str(patient[1]), style, 1)

                sum_price = 0
                inquiry = f"""SELECT DISTINCT records.id, lessons.id, places.price FROM places, records, lessons
                                    WHERE records.is_deleted = 0 AND lessons.is_deleted = 0 AND
    (lessons.id = records.lesson_id_1 OR lessons.id = records.lesson_id_2 OR lessons.id = records.lesson_id_3) AND
    records.patient_id = {patient[0]} AND lessons.id_plase = places.id AND lessons.id_doctor != 0 AND lessons.id_plase != 0"""
                prices = self.cur.execute(inquiry).fetchall()
                sum_lessons = len(prices)
                for price in prices:
                    sum_price += price[2]
                global_sum_lessons += sum_lessons
                global_sum_price += sum_price

                set_cell(sheet, num_str, 3, str(sum_lessons), style, 2)
                set_cell(sheet, num_str, 4, str(sum_price), style, 2)

            num_str += 1
            set_cell(sheet, num_str, 1, 'Всего', bold_style, 3)
            set_cell(sheet, num_str, 2, str(count_right_patients), bold_style, 2)
            set_cell(sheet, num_str, 3, str(global_sum_lessons), bold_style, 2)
            set_cell(sheet, num_str, 4, str(global_sum_price), bold_style, 2)
            num_str += 3

        num_str += 2
        set_cell(sheet, num_str, 1, f'* Данный отчёт составлен с помощью программы "Журнал ЛФК" версии {version}',
                 pale_style)

        sheet.column_dimensions['A'].width = 11
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 13
        sheet.column_dimensions['D'].width = 11

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        # sheet.orientation = 'landscape'
        # sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("Ошибка")

            msg.setText("Не удалось сохранить отчёт")
            msg.setInformativeText(
                "Файл который вы выбрали сейчас используется, поэтому в него невозможно записать отчёт.\n\nВыбрать новый файл для сохранения или попробовать ещё раз?")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.Retry)

            result = msg.exec_()

            if result == 16384:
                result = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
                if not result:
                    return
                file_name = result
            try:
                book.save(file_name)
            except PermissionError:
                pass
        self.close()  # закрывает это окно
        self.main_menu.show()

    def resizeEvent(self, event):
        pass

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()



if __name__ == '__main__':
    app = QApplication(sys.argv)  # создаёт и отображает окно входа в систему
    ex = ReportMenu6(None, '', 'C:/Users/pavel/Desktop/журнал ЛФК/lfk.db')
    ex.show()
    sys.exit(app.exec_())

