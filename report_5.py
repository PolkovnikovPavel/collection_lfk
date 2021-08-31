import os, sqlite3, openpyxl, datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox
from PyQt5.QtWidgets import QGridLayout, QProgressBar, QWidget, QCheckBox
from PyQt5 import QtGui, QtCore

from data.design.report_5 import Ui_MainWindow as Ui_Report5

bold_big_style = Font(size="14", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")

green_fill = PatternFill(start_color='FF66CC33', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFFFF33', fill_type='solid')
red_fill = PatternFill(start_color='FFFF3333', fill_type='solid')
purple_fill = PatternFill(start_color='FF9966CC', fill_type='solid')
colors = {1: green_fill, 2: yellow_fill, 3: red_fill, 4: purple_fill, 5: None}
white_blue = PatternFill(start_color='FFf2f8fe', fill_type='solid')
white_yellow = PatternFill(start_color='FFf5f5dc', fill_type='solid')
color_light_gray = PatternFill(start_color='e1e1e1', fill_type='solid')

border_top = Border(top=Side(style='thick'))
border_t_r_l = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     top=Side(style='thick'))
border_r_l_t_b = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     bottom=Side(style='medium'),
                     top=Side(style='medium'))
border_b_r_l = Border(left=Side(style='medium'),
                     right=Side(style='medium'),
                     bottom=Side(style='thick'))
border_r_l_thick = Border(left=Side(style='thick'),
                     right=Side(style='thick'))
border_r_l_b_thick = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     bottom=Side(style='thick'))

abc = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
       'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def set_cell(sheet, x, y, value, style, alignment=1, fill=None, border=None):
    cell = sheet._get_cell(x, y)
    cell.value = value
    cell.font = style
    if alignment == 1:
        cell.alignment = Alignment(horizontal='left')
    elif alignment == 2:
        cell.alignment = Alignment(horizontal='center')
    elif alignment == 3:
        cell.alignment = Alignment(horizontal='right')
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border


def get_datetime_from_text(text):
    text = text.split('.')
    datetime_date = datetime.date(day=int(text[0]), month=int(text[1]), year=int(text[2]))
    return datetime_date


def get_num_from_date(date):
    d, m, y = date.split('.')
    return int(m) * 32 + int(y) * 385 + int(d)


def get_date_calendar(calendar):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def get_datetime_date_from_calendar(calendar):
    calendar_date = calendar.selectedDate()
    datetime_date = datetime.date(day=calendar_date.day(), month=calendar_date.month(),
                         year=calendar_date.year())
    return datetime_date


class ReportMenu5(QMainWindow, Ui_Report5):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.button.clicked.connect(self.create_report)
        self.choice_date_1.clicked['QDate'].connect(self.choiced_date_1)
        self.choice_date_2.clicked['QDate'].connect(self.choiced_date_2)
        self.setWindowTitle('Сборный отчёт динамики лечения')

        layout = QGridLayout()  # создаёт настроеный layout
        inquiry = f"""SELECT DISTINCT id, name FROM places WHERE is_deleted = 0"""
        all_places = self.cur.execute(inquiry).fetchall()
        font = QtGui.QFont()
        font.setPointSize(14)
        self.right_places = []
        i = -1
        for place in all_places:
            i += 1
            cb = QCheckBox(str(place[1]), self)
            cb.args = place
            cb.stateChanged.connect(self.change_place)
            cb.setFont(font)
            layout.addWidget(cb, i, 0)  # в layout добавляются созданые элементы

        widget = QWidget()
        widget.setLayout(layout)  # устанавливает этот layout в widget
        self.scroll_places.setWidget(widget)


        layout = QGridLayout()  # создаёт настроеный layout
        inquiry = f"""SELECT DISTINCT id, name, short_name FROM accounts WHERE is_deleted = 0"""
        all_doctor = self.cur.execute(inquiry).fetchall()
        self.right_doctors = []
        i = -1
        for doctor in all_doctor:
            i += 1
            cb = QCheckBox(f'{doctor[1]} ({doctor[2]})', self)
            cb.args = doctor
            cb.stateChanged.connect(self.change_doctor)
            cb.setFont(font)
            layout.addWidget(cb, i, 0)  # в layout добавляются созданые элементы

        widget = QWidget()
        widget.setLayout(layout)  # устанавливает этот layout в widget
        self.scroll_doctor.setWidget(widget)

    def change_place(self, state):
        if state:
            self.right_places.append(self.sender().args[0])
        else:
            del self.right_places[self.right_places.index(self.sender().args[0])]

    def change_doctor(self, state):
        if state:
            self.right_doctors.append(self.sender().args[0])
        else:
            del self.right_doctors[self.right_doctors.index(self.sender().args[0])]

    def choiced_date_1(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_1)) >
           get_num_from_date(get_date_calendar(self.choice_date_2))):
            self.choice_date_2.setSelectedDate(self.choice_date_1.selectedDate())

    def choiced_date_2(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_2)) <
           get_num_from_date(get_date_calendar(self.choice_date_1))):
            self.choice_date_1.setSelectedDate(self.choice_date_2.selectedDate())

    def create_report(self):
        if len(self.right_places) == 0:
            self.error_text.setText('Выберите хотя бы один способ реабилитации')
            return
        if len(self.right_doctors) == 0:
            self.error_text.setText('Выберите хотя бы одного врача')
            return


        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return
        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == "сборный":
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        right_places = self.right_places
        right_doctors = self.right_doctors
        start_date = get_datetime_date_from_calendar(self.choice_date_1)
        end_date = get_datetime_date_from_calendar(self.choice_date_2)

        sheet.title = "сборный"

        set_cell(sheet, 1, 1, 'Сборный отчёт динамики лечения', bold_big_style)

        set_cell(sheet, 2, 1, 'ФИО, год рождения, диагноз', bold_style)
        set_cell(sheet, 2, 2, '№ карты', bold_style)
        set_cell(sheet, 2, 3, 'дата с..по..', bold_style)
        set_cell(sheet, 2, 4, 'дней пребывания', bold_style)
        set_cell(sheet, 2, 5, 'дней лечения', bold_style)
        set_cell(sheet, 2, 6, 'всего процедур', bold_style)
        set_cell(sheet, 2, 7, 'всего ед.', bold_style)

        inquiry = f"""SELECT DISTINCT * FROM patients
                            WHERE is_discharge = 1 and is_deleted = 0"""
        all_patients = self.cur.execute(inquiry).fetchall()
        inquiry = f"""SELECT DISTINCT * FROM places
                                    WHERE is_deleted = 0"""
        _all_places = self.cur.execute(inquiry).fetchall()
        all_places = []
        for place in _all_places:
            if place[0] in right_places:
                all_places.append(place)
        num_str = 3
        color_patient = white_blue
        color_record = white_blue
        for patient in all_patients:
            date_of_operation = get_datetime_from_text(patient[10])
            if date_of_operation < start_date or date_of_operation > end_date:
                continue
            delte_datetime = get_datetime_from_text(patient[9]) - date_of_operation
            set_cell(sheet, num_str, 1, f'{patient[1]}, {patient[2]},', style)
            set_cell(sheet, num_str + 1, 1, 'диагноз:', style)
            set_cell(sheet, num_str + 2, 1, patient[6], style)
            set_cell(sheet, num_str, 2, patient[3], style)
            set_cell(sheet, num_str, 3, f'c {patient[10]} по {patient[9]}', style)
            set_cell(sheet, num_str, 4, delte_datetime.days, style, 2)

            rows = sheet.iter_rows(num_str, num_str, 1, 50)
            for row in rows:
                for cell in row:
                    cell.border = border_top

            if color_patient == white_blue:
                color_patient = white_yellow
            else:
                color_patient = white_blue
            rows = sheet.iter_rows(num_str, num_str + 3 * len(all_places), 1, 4)
            for row in rows:
                for cell in row:
                    cell.fill = color_patient

            num_place = -1
            days_treatment = 0
            count_records = 0
            total_price = 0
            for i in range(len(all_places)):
                num_place += 1
                place = all_places[i]
                if color_record == white_blue:
                    color_record = white_yellow
                else:
                    color_record = white_blue
                rows = sheet.iter_rows(num_str + num_place * 3, num_str + num_place * 3 + 3, 5, 50)
                for row in rows:
                    for cell in row:
                        cell.fill = color_record

                inquiry = f"""SELECT DISTINCT records.id, records.date, places.price, places.short_name,
records.grade_1, records.grade_2, records.grade_3, accounts.short_name, accounts.id FROM records, lessons, places, accounts
                                WHERE records.patient_id = {patient[0]} and records.is_deleted = 0
and lessons.id_plase = {place[0]} and lessons.id_doctor != 0 and 
(lessons.id = records.lesson_id_1 or lessons.id = records.lesson_id_2 or lessons.id = records.lesson_id_3) and
places.id = {place[0]} and accounts.id = lessons.id_doctor"""
                all_records = self.cur.execute(inquiry).fetchall()
                different_dates = []
                for record in all_records:
                    if record[1] not in different_dates:
                        different_dates.append(record[1])
                price = place[3]

                days_treatment += len(different_dates)
                count_records += len(all_records)
                total_price += len(all_records) * price
                set_cell(sheet, num_str + num_place * 3 + 1, 5, f'{place[4]} -  {len(different_dates)}', style)
                set_cell(sheet, num_str + num_place * 3 + 1, 6, f'{len(all_records)}', style, 2)
                set_cell(sheet, num_str + num_place * 3 + 1, 7, f'* {price} = {len(all_records) * price}', style)

                calum_num = 7
                for record in all_records:
                    grade_1, grade_2, grade_3 = record[4], record[5], record[6]
                    doctor = record[7]
                    if grade_1 == 5 and grade_2 == 5 and grade_3 == 5:
                        continue
                    if record[8] not in right_doctors:
                        continue
                    calum_num += 1

                    set_cell(sheet, num_str + num_place * 3 + 0, calum_num, '', style, fill=colors[grade_1], border=border_t_r_l)
                    set_cell(sheet, num_str + num_place * 3 + 1, calum_num, str(doctor), style, 2, fill=colors[grade_2], border=border_r_l_t_b)
                    set_cell(sheet, num_str + num_place * 3 + 2, calum_num, '', style, fill=colors[grade_3], border=border_b_r_l)

            set_cell(sheet, num_str + num_place * 3 + 3, 5, f'всего: {days_treatment}', style)
            set_cell(sheet, num_str + num_place * 3 + 3, 6, f'{count_records}', style, 2)
            set_cell(sheet, num_str + num_place * 3 + 3, 7, f'{total_price}', style, 2)

            num_str += len(right_places) * 3 + 1

        rows = sheet.iter_rows(2, num_str, 5, 5)
        for row in rows:
            for cell in row:
                if (cell.row - 2) % (len(right_places) * 3 + 1) == 0:
                    cell.border = border_r_l_b_thick
                else:
                    cell.border = border_r_l_thick
        rows = sheet.iter_rows(2, num_str, 7, 7)
        for row in rows:
            for cell in row:
                if (cell.row - 2) % (len(right_places) * 3 + 1) == 0:
                    cell.border = border_r_l_b_thick
                else:
                    cell.border = border_r_l_thick

        for symbol_1 in abc[0:5]:
            for symbol_2 in abc[1:]:
                sheet.column_dimensions[symbol_1 + symbol_2].width = 4.2
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 23
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 12.9
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 8

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.orientation = 'landscape'
        sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Question)
            msg.setWindowTitle("Ошибка")

            msg.setText("Не удалось сохранить отчёт")
            msg.setInformativeText(
                "Файл который вы выбрали сейчас используется, поэтому в него невозможно записать отчёт.\n\nВыбрать новый файл для сохранения или попробовать ещё раз?")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.Retry)

            result = msg.exec_()

            if result == 16384:
                result = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
                if not result:
                    return
                file_name = result
            try:
                book.save(file_name)
            except PermissionError:
                pass
        self.close()  # закрывает это окно
        self.main_menu.show()

    def resizeEvent(self, event):
        self.button.setGeometry(QtCore.QRect(self.width() - 225, self.height() - 65, 200, 40))
        self.scroll_doctor.setGeometry(QtCore.QRect(self.width() // 2, 480, self.width() // 2 - 20, self.height() - 580))
        self.label_6.setGeometry(QtCore.QRect(self.width() // 2, 450, 470, 30))
        self.scroll_places.setGeometry(QtCore.QRect(30, 480, self.width() // 2 - 150, self.height() - 580))
        self.choice_date_2.setGeometry(QtCore.QRect(self.width() - 460, 120, 450, 310))
        self.label_3.setGeometry(QtCore.QRect(self.width() - 420, 80, 450, 31))
        self.error_text.setGeometry(QtCore.QRect(30, self.height() - 60, 700, 30))

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
