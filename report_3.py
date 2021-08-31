import os, sqlite3, openpyxl
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QInputDialog

from data.design.report_3 import Ui_MainWindow as Ui_Report3
from report_5 import set_cell, color_light_gray
from report_2 import get_detailed_report_month, border_t_b_bold, border_r_b, border_bottom, border_right


def get_date_calendar(calendar):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


all_month = {1: 'янв.', 2: 'февр.', 3: 'март', 4: 'апр.',
             5: 'май', 6: 'июнь', 7: 'июль', 8: 'авг.',
             9: 'сент.', 10: 'окт.', 11: 'нояб.', 12: 'дек.'}

all_month_full = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
             '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
             '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}

month_convert = {1: '01', 2: '02', 3: '03', 4: '04',
             5: '05', 6: '06', 7: '07', 8: '08',
             9: '09', 10: '10', 11: '11', 12: '12'}


class ReportMenu3(QMainWindow, Ui_Report3):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.setWindowTitle('Отчёт за год')
        self.button.clicked.connect(self.create_report)
        self.name_year.activated.connect(self.select_year)

        inquiry = f"""SELECT DISTINCT date FROM records
                                            WHERE is_deleted = 0"""
        all_all_dates = self.cur.execute(inquiry).fetchall()
        self.all_years = []
        for date in all_all_dates:
            date = int(date[0].split('.')[2])
            if date not in self.all_years:
                self.all_years.append(date)

        self.all_years.sort(reverse=True)

        self.selected_year = self.all_years[0]
        for year in self.all_years:
            self.name_year.addItem(str(year))

    def select_year(self):
        for month in self.all_years:
            if self.name_year.currentText() == str(month):
                self.selected_year = month


    def create_report(self):
        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return

        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == f"отчёт за {self.selected_year} год":
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        sheet.title = f"отчёт за {self.selected_year} год"

        bold_big_style = Font(size="14", bold=True)
        bold_style = Font(size="11", bold=True)
        style = Font(size="11")
        pale_style = Font(size="9", color='00555555', italic=True)

        # The data
        set_cell(sheet, 1, 1, 'Отчет за год работы кабинета реабилитации', bold_big_style)
        set_cell(sheet, 1, 9, str(self.selected_year) + 'г.', bold_big_style)


        inquiry = f"""SELECT DISTINCT name FROM categories WHERE is_deleted = 0"""
        all_categories = self.cur.execute(inquiry).fetchall()

        inquiry = f"""SELECT DISTINCT name FROM departments WHERE is_deleted = 0"""
        all_departments = self.cur.execute(inquiry).fetchall()

        inquiry = f"""SELECT DISTINCT id, name, short_name, is_deleted FROM accounts"""
        all_doctors = self.cur.execute(inquiry).fetchall()

        # ---------------------------------------катигории людей и их отделения
        inquiry = f"""SELECT records.id, patients.id FROM records, patients
                        WHERE records.is_deleted = 0 and date LIKE '%{self.selected_year}'
                                        and records.patient_id = patients.id and patients.is_deleted = 0"""
        all_records = self.cur.execute(inquiry).fetchall()

        all_people_in_year = []
        right_id_people = []

        for record in all_records:
            if record[1] not in all_people_in_year:
                all_people_in_year.append(record[1])

        for id_people in all_people_in_year:
            inquiry = f"""SELECT records.id, records.date FROM records
            WHERE records.is_deleted = 0 and records.patient_id = {id_people}"""
            all_records = self.cur.execute(inquiry).fetchall()
            is_right_people = True
            min_month = 12
            for record in all_records:
                date = record[1].split('.')
                if int(date[2]) < self.selected_year:
                    is_right_people = False
                    break
                if int(date[1]) < min_month:
                    min_month = int(date[1])
            if is_right_people:
                right_id_people.append([id_people, min_month])

        department_by_people = {}
        for department in all_departments:
            department_by_people[department[0]] = [0 for _ in range(12)]
        department_by_people['удал. отделения'] = [0 for _ in range(12)]

        category_by_people = {}
        for categorie in all_categories:
            category_by_people[categorie[0]] = [0 for _ in range(12)]
        category_by_people['удал. категории'] = [0 for _ in range(12)]

        for people in right_id_people:
            id_people, min_month = people
            inquiry = f"""SELECT patients.id, departments.name, categories.name FROM patients, departments, categories
WHERE patients.id = {id_people} and patients.category = categories.id and patients.department = departments.id and patients.is_deleted = 0"""
            department_categori = self.cur.execute(inquiry).fetchone()
            if department_categori[1] not in department_by_people:
                department_by_people['удал. отделения'][min_month - 1] += 1
            else:
                department_by_people[department_categori[1]][min_month - 1] += 1
            if department_categori[2] not in category_by_people:
                category_by_people['удал. категории'][min_month - 1] += 1
            else:
                category_by_people[department_categori[2]][min_month - 1] += 1


        set_cell(sheet, 3, 4, 'год', bold_style, 2, border=border_r_b)
        for i in range(1, 13):
            set_cell(sheet, 3, 4 + i, all_month[i], style, 2, border=border_bottom)
            set_cell(sheet, 4, 4 + i, len(list(filter(lambda x: x[1] == i, right_id_people))), bold_style, 2)
        set_cell(sheet, 4, 1, 'Всего человек', bold_style)
        set_cell(sheet, 4, 4, len(right_id_people), bold_style, 2, border=border_right)
        if sum(category_by_people['удал. категории']) != 0:
            all_categories.append(['удал. категории'])
        num_str = 4
        for categorie in all_categories:
            num_str += 1
            set_cell(sheet, num_str, 1, categorie[0], style)
            set_cell(sheet, num_str, 4, sum(category_by_people[categorie[0]]), bold_style, 2, border=border_right)
            for i in range(12):
                set_cell(sheet, num_str, 5 + i, category_by_people[categorie[0]][i], style, 2)


        set_cell(sheet, num_str + 2, 1, 'По отделениям', bold_style, border=border_right)
        if sum(department_by_people['удал. отделения']) != 0:
            all_departments.append(['удал. отделения'])

        set_cell(sheet, num_str + 1, 4, '', style, border=border_right)
        set_cell(sheet, num_str + 2, 4, '', style, border=border_right)
        num_str += 2
        for department in all_departments:
            num_str += 1
            set_cell(sheet, num_str, 1, department[0], style)
            set_cell(sheet, num_str, 4, sum(department_by_people[department[0]]), bold_style, 2, border=border_right)
            for i in range(12):
                set_cell(sheet, num_str, 5 + i, department_by_people[department[0]][i], style, 2)
        # ---------------------------------------------------------------------

        # ------------------------------------------------------- По процедурам

        text = f'Способ реабилитации за {self.selected_year} Год'
        year = get_detailed_report_month(self.selected_year, self.cur, False, text, mod=1, is_year=True, all_doctors=all_doctors)
        months = []
        for i in range(1, 13):
            date = month_convert[i]
            text = f'Способ реабилитации за {all_month_full[date]}'
            month = get_detailed_report_month(date, self.cur, True, text, mod=1, all_doctors=all_doctors)
            months.append(month)

        set_cell(sheet, num_str + 1, 4, '', style, border=border_right)
        num_str += 2
        set_cell(sheet, num_str, 1, 'По процедурам', bold_style)
        set_cell(sheet, num_str + 1, 1, 'по единицам ', bold_style)

        set_cell(sheet, num_str, 4, year[1], bold_style, 2, border=border_right)
        set_cell(sheet, num_str + 1, 4, year[2], bold_style, 2, border=border_right)
        for i in range(12):
            set_cell(sheet, num_str, 5 + i, months[i][1], style, 2)
            set_cell(sheet, num_str + 1, 5 + i, months[i][2], style, 2)

        num_str += 3
        set_cell(sheet, num_str, 1, '* данные ниже - о количестве человек, рассчитаны другим методом в отличие от верхней части отчёта и поэтому не стоит их брать в расчёт (показывают количество разных людей за данный период)', pale_style)
        lines = year[0]
        for line in lines:
            num_str += 1
            for i in range(1, len(line) + 1):
                call = line[i - 1]
                set_cell(sheet, num_str, i, call.text, call.style, border=call.border, alignment=call.alignment)

        num_str += 2
        for month in months:
            lines = month[0]
            for line in lines:
                num_str += 1
                for i in range(1, len(line) + 1):
                    call = line[i - 1]
                    set_cell(sheet, num_str, i, call.text, call.style, border=call.border, alignment=call.alignment)
            num_str += 1
            for i in range(1, len(line) + 1):
                set_cell(sheet, num_str, i, '', style, border=border_t_b_bold, fill=color_light_gray)

        # ______________________________________________________________

        num_str += 3
        set_cell(sheet, num_str, 2, 'Исполнители', bold_style)
        for doctor in all_doctors:
            if doctor[3]:
                continue
            num_str += 1
            set_cell(sheet, num_str, 1, doctor[1] + '…' * 40, style)
            set_cell(sheet, num_str, 6, doctor[2], bold_style)
        num_str += 1
        set_cell(sheet, num_str, 2, 'Удалённые исполнители', bold_style)
        for doctor in all_doctors:
            if doctor[3]:
                num_str += 1
                set_cell(sheet, num_str, 1, doctor[1] + '…' * 40, style)
                set_cell(sheet, num_str, 6, doctor[2], bold_style)


        sheet.column_dimensions['E'].width = 5.5
        sheet.column_dimensions['F'].width = 6

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.orientation = 'landscape'
        sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            pass
        self.close()

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
