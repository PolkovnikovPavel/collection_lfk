import os, sqlite3, openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QCalendarWidget
from PyQt5 import QtCore, QtGui, QtWidgets
from report_1 import create_report_on_day
from datetime import date
from dateutil.relativedelta import relativedelta

from data.design.report_4 import Ui_MainWindow as Ui_Report4

bold_big_style = Font(size="14", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")


def set_cell(sheet, x, y, value, style, alignment=1):
    cell = sheet._get_cell(x, y)
    cell.value = value
    cell.font = style
    if alignment == 1:
        cell.alignment = Alignment(horizontal='left')
    elif alignment == 2:
        cell.alignment = Alignment(horizontal='center')
    elif alignment == 3:
        cell.alignment = Alignment(horizontal='right')


def get_num_from_date(date):
    d, m, y = date.split('.')
    return int(m) * 32 + int(y) * 385 + int(d)


def get_date_calendar(calendar, mod=0):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    if mod == 0:
        date = calendar.selectedDate()
        date = f'{date.day()}.{date.month()}.{date.year()}'
    else:
        date = f'{calendar.day}.{calendar.month}.{calendar.year}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def get_datetime_date_from_calendar(calendar):
    calendar_date = calendar.selectedDate()
    datetime_date = date(day=calendar_date.day(), month=calendar_date.month(),
                         year=calendar_date.year())
    return datetime_date


class ReportMenu4(QMainWindow, Ui_Report4):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.button.clicked.connect(self.create_report)
        self.choice_date_1.clicked['QDate'].connect(self.choiced_date_1)
        self.choice_date_2.clicked['QDate'].connect(self.choiced_date_2)

    def choiced_date_1(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_1)) >
           get_num_from_date(get_date_calendar(self.choice_date_2))):
            self.choice_date_2.setSelectedDate(self.choice_date_1.selectedDate())

    def choiced_date_2(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_2)) <
           get_num_from_date(get_date_calendar(self.choice_date_1))):
            self.choice_date_1.setSelectedDate(self.choice_date_2.selectedDate())

    def create_report(self):

        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return

        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == "отчёт за период дней":
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        sheet.title = "отчёт за период дней"

        num_str = 1
        date_1 = get_datetime_date_from_calendar(self.choice_date_1)
        date_2 = get_datetime_date_from_calendar(self.choice_date_2)
        while True:
            num_str = create_report_on_day(sheet, self.cur, get_date_calendar(date_1, 1), num_str)

            page_break = Break(id=num_str - 1)
            sheet.row_breaks.append(page_break)  # insert page break

            if date_1.day == date_2.day and date_1.month == date_2.month and date_1.year == date_2.year:
                break
            date_1 = date_1 + relativedelta(days=1)

        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['F'].width = 7
        sheet.column_dimensions['H'].width = 7
        sheet.column_dimensions['J'].width = 7
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['E'].width = 24

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.orientation = 'landscape'
        sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            pass
        self.close()

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
