import os, sqlite3, openpyxl
from openpyxl.styles import Font, Alignment
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QInputDialog
from PyQt5 import QtCore, QtGui, QtWidgets

from data.design.report_1 import Ui_MainWindow as Ui_Report1

bold_big_style = Font(size="14", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")


def set_cell(sheet, x, y, value, style, alignment=1):
    cell = sheet._get_cell(x, y)
    cell.value = value
    cell.font = style
    if alignment == 1:
        cell.alignment = Alignment(horizontal='left')
    elif alignment == 2:
        cell.alignment = Alignment(horizontal='center')
    elif alignment == 3:
        cell.alignment = Alignment(horizontal='right')


def get_date_calendar(calendar):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def create_report_on_day(sheet, cur, choice_date, start_num_str):
    set_cell(sheet, start_num_str, 1, 'Журнал работы с пациентами', bold_big_style)

    set_cell(sheet, start_num_str, 7, choice_date, bold_big_style)

    set_cell(sheet, start_num_str + 3, 1, '№', bold_style)
    set_cell(sheet, start_num_str + 3, 2, 'ФИО, год рождения, диагноз', bold_style)
    set_cell(sheet, start_num_str + 3, 3, '№ истории', bold_style)
    set_cell(sheet, start_num_str + 3, 4, 'взр./реб.', bold_style)
    set_cell(sheet, start_num_str + 3, 5, 'Отдел', bold_style)
    set_cell(sheet, start_num_str + 3, 6, 'процед.', bold_style)
    set_cell(sheet, start_num_str + 3, 7, 'исп.', bold_style)
    set_cell(sheet, start_num_str + 3, 8, 'процед.', bold_style)
    set_cell(sheet, start_num_str + 3, 9, 'исп.', bold_style)
    set_cell(sheet, start_num_str + 3, 10, 'процед.', bold_style)
    set_cell(sheet, start_num_str + 3, 11, 'исп.', bold_style)

    num_str = start_num_str + 4

    inquiry = f"""SELECT DISTINCT patient_id, lesson_id_1, lesson_id_2, lesson_id_3
                                                            FROM records, patients
                            WHERE date = '{choice_date}' AND records.is_deleted = 0 
                            and patients.is_deleted = 0 and patients.id = records.patient_id"""
    all_records = cur.execute(inquiry).fetchall()

    count_of_patients = {}
    count_of_records = 0
    all_places = {}

    data = []
    for record in all_records:
        if record[0] not in count_of_patients:
            count_of_patients[record[0]] = 1
        else:
            count_of_patients[record[0]] += 1
        inquiry = f"""SELECT DISTINCT full_name, diagnosis, date_of_birth, story_number, category, department, my_story_number
                                                    FROM patients
                                                    WHERE id = {record[0]}"""
        patient = cur.execute(inquiry).fetchone()

        stitch = []
        stitch.append(patient[6])
        stitch.append(patient[0] + ' ' + patient[2] + ' ' + patient[1])
        stitch.append(patient[3])

        inquiry = f"SELECT DISTINCT name FROM categories WHERE id = {patient[4]}"
        categorie = cur.execute(inquiry).fetchone()
        stitch.append(categorie[0])

        inquiry = f"SELECT DISTINCT name FROM departments WHERE id = {patient[5]}"
        department = cur.execute(inquiry).fetchone()
        stitch.append(department[0])

        inquiry = f"SELECT DISTINCT * FROM lessons WHERE id = {record[1]}"
        lesson_1 = cur.execute(inquiry).fetchone()
        inquiry = f"SELECT DISTINCT * FROM lessons WHERE id = {record[2]}"
        lesson_2 = cur.execute(inquiry).fetchone()
        inquiry = f"SELECT DISTINCT * FROM lessons WHERE id = {record[3]}"
        lesson_3 = cur.execute(inquiry).fetchone()

        for lesson in (lesson_1, lesson_2, lesson_3):
            if lesson[3] == 0 and lesson[1] != 0 and lesson[2] != 0:
                count_of_records += 1
                inquiry = f"SELECT DISTINCT name FROM places WHERE id = {lesson[1]}"
                place = cur.execute(inquiry).fetchone()
                stitch.append(place[0])
                if place[0] not in all_places:
                    all_places[place[0]] = 1
                else:
                    all_places[place[0]] += 1

                inquiry = f"SELECT DISTINCT short_name FROM accounts WHERE id = {lesson[2]}"
                name = cur.execute(inquiry).fetchone()
                stitch.append(name[0])
            else:
                stitch.append('---------')
                stitch.append('-------------')

        data.append(stitch)

    for i in range(len(all_records)):
        set_cell(sheet, num_str, 1, data[i][0], style)
        set_cell(sheet, num_str, 2, data[i][1], style)
        set_cell(sheet, num_str, 3, data[i][2], style)
        set_cell(sheet, num_str, 4, data[i][3], style)
        set_cell(sheet, num_str, 5, data[i][4], style)
        set_cell(sheet, num_str, 6, data[i][5], style)
        set_cell(sheet, num_str, 7, data[i][6], style)
        set_cell(sheet, num_str, 8, data[i][7], style)
        set_cell(sheet, num_str, 9, data[i][8], style)
        set_cell(sheet, num_str, 10, data[i][9], style)
        set_cell(sheet, num_str, 11, data[i][10], style)
        num_str += 1

    set_cell(sheet, num_str + 2, 1, f'Всего процедур за {choice_date}:', bold_style)
    set_cell(sheet, num_str + 2, 3, count_of_records, bold_style)
    set_cell(sheet, num_str + 2, 4, 'Всего пациентов :', bold_style)
    set_cell(sheet, num_str + 2, 5, len(count_of_patients), bold_style)

    key_places = list(all_places.keys())
    key_places.sort()
    num_str += 3
    for name_place in key_places:
        points = '…' * (50 - len(name_place))
        set_cell(sheet, num_str, 1, name_place + points, style)
        set_cell(sheet, num_str, 3, all_places[name_place], bold_style)
        num_str += 1

    return num_str


class ReportMenu1(QMainWindow, Ui_Report1):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.button.clicked.connect(self.create_report)

    def create_report(self):

        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return

        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == "отчёт за день":
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        sheet.title = "отчёт за день"

        create_report_on_day(sheet, self.cur, get_date_calendar(self.choice_date), 1)

        sheet.column_dimensions['B'].width = 55
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 17
        sheet.column_dimensions['F'].width = 7
        sheet.column_dimensions['H'].width = 7
        sheet.column_dimensions['J'].width = 7
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['E'].width = 24

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.orientation = 'landscape'
        sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            pass
        self.close()

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
