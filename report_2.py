import os, sqlite3, openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QInputDialog
from PyQt5 import QtCore, QtGui
from datetime import date
from dateutil.relativedelta import relativedelta

from data.design.report_2 import Ui_MainWindow as Ui_Report2
from report_5 import set_cell, color_light_gray
from description_menu import version


all_month = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
             '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
             '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}

bold_big_style = Font(size="14", bold=True)
bold_medium_style = Font(size="13", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")
pale_style = Font(size="9", color='00555555', italic=True)


border_t_b = Border(bottom=Side(style='medium'),
                     top=Side(style='medium'))
border_left = Border(left=Side(style='medium'))
border_t_b_bold = Border(bottom=Side(style='thick'),
                     top=Side(style='thick'))
border_r_b = Border(bottom=Side(style='medium'),
                     right=Side(style='medium'))
border_bottom = Border(bottom=Side(style='medium'))
border_right = Border(right=Side(style='medium'))

class Call:
    def __init__(self, text, style=style, border=None, alignment=1):
        self.text = text
        self.style = style
        self.border = border
        self.alignment = alignment



def get_date_calendar(calendar):
    # возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def get_datetime_date(day, month, year):
    datetime_date = date(day=day, month=month,
                         year=year)
    return datetime_date


def get_num_from_date(date):
    d, m, y = date.split('.')
    return int(m) * 31 + int(y) * 385 + int(d)


def get_detailed_report_month(date, cur, is_month=True, main_text='Способ реабилитации', all_doctors=None, mod=0, is_year=False):
    inquiry = f"""SELECT DISTINCT id, name, price FROM places WHERE is_deleted = 0"""
    all_places = cur.execute(inquiry).fetchall()
    inquiry = f"""SELECT DISTINCT id, name FROM categories WHERE is_deleted = 0"""
    all_categories = cur.execute(inquiry).fetchall()


    if is_month:
        date = f'___{date[0]}.{date[1]}'
    elif is_year:
        date = f'%{date}'

    inquiry = f"""SELECT records.id, places.id, places.name, accounts.id, accounts.name, patients.id, categories.name, places.price FROM records, lessons, places, accounts, patients, categories
                WHERE records.date LIKE '{date}' and (lessons.id = records.lesson_id_1 or 
                                                      lessons.id = records.lesson_id_2 or
                                                      lessons.id = records.lesson_id_3) and categories.id = patients.category and
                lessons.id_plase = places.id and lessons.id_doctor = accounts.id and records.is_deleted = 0 and patients.is_deleted = 0 and patients.id = records.patient_id"""
    all_records = cur.execute(inquiry).fetchall()

    records_by_places = {}
    doctors_by_places = {}
    right_id_people = []
    right_id_people_by_doctors = {}

    right_id_people_by_records = {}
    right_id_people_by_records_by_doc = {}



    for place in all_places:
        records_by_places[place[0]] = 0
        right_id_people_by_records[place[0]] = set()
        right_id_people_by_records_by_doc[place[0]] = {}
    records_by_places['удал. места'] = 0
    right_id_people_by_records['удал. места'] = set()
    right_id_people_by_records_by_doc['удал. места'] = {}

    if all_doctors is None:
        inquiry = f"""SELECT DISTINCT id, name, short_name, is_deleted FROM accounts"""
        all_doctors = cur.execute(inquiry).fetchall()
    for doctor in all_doctors:
        vocabulary = {}
        for i in records_by_places:
            vocabulary[i] = 0
            right_id_people_by_records_by_doc[i][doctor[0]] = set()
        doctors_by_places[doctor[0]] = vocabulary
        right_id_people_by_doctors[doctor[0]] = []

    records_by_category = {}
    for category in all_categories:
        records_by_category[category[1]] = [0, 0]

    for record in all_records:
        if record[1] not in records_by_places:
            records_by_places['удал. места'] += 1
            doctors_by_places[record[3]]['удал. места'] += 1
            right_id_people_by_records['удал. места'].add(record[5])
            right_id_people_by_records_by_doc['удал. места'][record[3]].add(record[5])
        else:
            records_by_places[record[1]] += 1
            doctors_by_places[record[3]][record[1]] += 1
            right_id_people_by_records[record[1]].add(record[5])
            right_id_people_by_records_by_doc[record[1]][record[3]].add(record[5])

        if record[5] not in right_id_people:
            right_id_people.append(record[5])
        if record[5] not in right_id_people_by_doctors[record[3]]:
            right_id_people_by_doctors[record[3]].append(record[5])
        if record[6] in records_by_category:
            records_by_category[record[6]][0] += 1
            records_by_category[record[6]][1] += record[7]
        else:
            if 'удал. категории' not in records_by_category:
                records_by_category['удал. категории'] = [1, record[7]]
            else:
                records_by_category['удал. категории'][0] += 1
                records_by_category['удал. категории'][1] += record[7]



    lines = []
    line = [Call(main_text, bold_style), Call(''), Call(''), Call(''), Call('чел.', bold_style, border_left), Call('пр.', bold_style), Call('еденицы', bold_style), Call('')]
    line2 = [Call(''), Call(''), Call(''), Call(''), Call('', style, border_left), Call(''), Call(''), Call('')]
    del_doctors = []
    for doctor in all_doctors:
        if doctor[3]:
            del_doctors.append(doctor)
            continue
        line.append(Call('', border=border_left))
        line.append(Call(doctor[2], bold_medium_style, alignment=2))
        line.append(Call(''))
        line2.extend([Call('чел.', border=border_left, alignment=2), Call('пр.', alignment=2), Call('ед.', alignment=2)])
    if len(del_doctors) > 0:
        line.append(Call('', border=border_left))
        line.append(Call('Удал. исполнители', alignment=border_left))
        line.append(Call(''))
        line2.extend([Call('чел.', border=border_left, alignment=2), Call('пр.', alignment=2), Call('ед.', alignment=2)])

    lines.append(line)
    lines.append(line2)
    sum_1 = 0
    sum_2 = 0

    if len(del_doctors) > 0:
        sums = [0 for _ in range((len(all_doctors) - len(del_doctors) + 1) * 2)]
    else:
        sums = [0 for _ in range(len(all_doctors) * 2)]
    for place in all_places:
        summ = records_by_places[place[0]] * place[2]
        line = [Call(place[1]), Call(''), Call(''), Call(''), Call(len(right_id_people_by_records[place[0]]), style, border_left, 3),
                Call(records_by_places[place[0]], style, alignment=3),
                Call(f'* {place[2]} =', style, border_left, 2), Call(summ)]
        sum_1 += records_by_places[place[0]]
        sum_2 += summ
        i = 0
        for doctor in all_doctors:
            if doctor[3]:
                continue
            human = len(right_id_people_by_records_by_doc[place[0]][doctor[0]])
            count = doctors_by_places[doctor[0]][place[0]]
            prise = count * place[2]
            if count == 0:
                count = '-'
                prise = '-'
                human = '-'

            line.append(Call(human, border=border_left, alignment=2))
            line.append(Call(count, alignment=2))
            line.append(Call(prise, alignment=2))
            sums[i] += doctors_by_places[doctor[0]][place[0]]
            sums[i + 1] += doctors_by_places[doctor[0]][place[0]] * place[2]
            i += 2
        del_doctors_human = 0
        del_doctors_doctor = 0
        del_doctors_prise = 0
        del_doctors_sum_1 = 0
        del_doctors_sum_2 = 0
        for doctor in del_doctors:
            del_doctors_human += len(right_id_people_by_records_by_doc[place[0]][doctor[0]])
            del_doctors_doctor += doctors_by_places[doctor[0]][place[0]]
            del_doctors_prise += doctors_by_places[doctor[0]][place[0]] * place[2]
            del_doctors_sum_1 += doctors_by_places[doctor[0]][place[0]]
            del_doctors_sum_2 += doctors_by_places[doctor[0]][place[0]] * place[2]
        if len(del_doctors) > 0:
            if del_doctors_human == 0:
                del_doctors_human = '-'
            if del_doctors_doctor == 0:
                del_doctors_doctor = '-'
            if del_doctors_prise == 0:
                del_doctors_prise = '-'
            line.append(Call(del_doctors_human, border=border_left, alignment=2))
            line.append(Call(del_doctors_doctor, alignment=2))
            line.append(Call(del_doctors_prise, alignment=2))
            sums[i] += del_doctors_sum_1
            sums[i + 1] += del_doctors_sum_2

        lines.append(line)

    line = [Call(''), Call(''), Call('всего:', bold_style), Call(''), Call(len(right_id_people), bold_style, border_left, 3),
            Call(sum_1, bold_style, None, 3), Call('', style, border_left), Call(sum_2, bold_style)]

    del_doctor_count = 0
    j = 0
    for i in right_id_people_by_doctors:
        if any(map(lambda x: x[0] == i, del_doctors)):
            del_doctor_count += len(right_id_people_by_doctors[i])
            continue
        line.append(Call(len(right_id_people_by_doctors[i]), bold_style, border_left, 2))
        line.append(Call(sums[j], bold_style, alignment=2))
        line.append(Call(sums[j + 1], bold_style, alignment=2))
        j += 2
    if len(del_doctors) != 0:
        line.append(Call(del_doctor_count, bold_style, border_left, 2))
        line.append(Call(sums[j], bold_style, alignment=2))
        line.append(Call(sums[j + 1], bold_style, alignment=2))
    lines.append(line)


    line = []
    for _ in range(8 + 3 * (len(sums) // 2)):
        line.append(Call('', style, border_t_b))
    lines.append(line)

    if mod == 0:
        return lines
    elif mod == 1:
        return lines, sum_1, sum_2, records_by_category




class ReportMenu2(QMainWindow, Ui_Report2):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.setWindowTitle('Отчёт за месяц')
        self.button.clicked.connect(self.create_report)
        self.name_month.activated.connect(self.select_month)

        inquiry = f"""SELECT DISTINCT date FROM records
                                            WHERE is_deleted = 0"""
        all_all_dates = self.cur.execute(inquiry).fetchall()
        self.all_dates = []
        for date in all_all_dates:
            date = '.'.join(date[0].split('.')[1::])
            num_date = date.split('.')
            date = all_month[date.split('.')[0]] + '(' + date.split('.')[0] + ') ' + date.split('.')[1]
            if (date, num_date) not in self.all_dates:
                self.all_dates.append((date, num_date))

        self.all_dates.sort(key=lambda x: int(x[1][0]) + int(x[1][1] * 12), reverse=True)

        self.selected_month = self.all_dates[0][1]
        for month in self.all_dates:
            self.name_month.addItem(month[0])


    def select_month(self):
        for month in self.all_dates:
            if self.name_month.currentText() == month[0]:
                self.selected_month = month[1]


    def create_report(self):
        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return

        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == f"отчёт за {self.name_month.currentText()}":
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        sheet.title = f"отчёт за {self.name_month.currentText()}"

        # The data
        set_cell(sheet, 1, 1, 'Отчет за месяц работы кабинета реабилитации', bold_big_style)
        set_cell(sheet, 1, 8, all_month[self.selected_month[0]] + ' ' + self.selected_month[1] + 'г.', bold_big_style)


        inquiry = f"""SELECT DISTINCT name FROM categories WHERE is_deleted = 0"""
        all_categories = self.cur.execute(inquiry).fetchall()

        inquiry = f"""SELECT DISTINCT name FROM departments WHERE is_deleted = 0"""
        all_departments = self.cur.execute(inquiry).fetchall()

        inquiry = f"""SELECT DISTINCT id, name, short_name, is_deleted FROM accounts"""
        all_doctors = self.cur.execute(inquiry).fetchall()

        # ---------------------------------------катигории людей и их отделения
        inquiry = f"""SELECT records.id, patients.id FROM records, patients
                        WHERE records.is_deleted = 0 and date LIKE '___{self.selected_month[0]}_____'
                                        and records.patient_id = patients.id and patients.is_deleted = 0"""
        all_records = self.cur.execute(inquiry).fetchall()

        all_people_in_month = []
        right_id_people = []

        for record in all_records:
            if record[1] not in all_people_in_month:
                all_people_in_month.append(record[1])

        for id_people in all_people_in_month:
            inquiry = f"""SELECT records.id, records.date FROM records
            WHERE records.is_deleted = 0 and records.patient_id = {id_people}"""
            all_records = self.cur.execute(inquiry).fetchall()
            if all(map(lambda x:
                       get_num_from_date(x[1]) >= get_num_from_date(f'01.{self.selected_month[0]}.{self.selected_month[1]}'),
                       all_records)):
                right_id_people.append(id_people)

        department_by_people = {}
        for department in all_departments:
            department_by_people[department[0]] = 0
        department_by_people['удал. отделения'] = 0

        category_by_people = {}
        for categorie in all_categories:
            category_by_people[categorie[0]] = 0
        category_by_people['удал. категории'] = 0

        for id_people in right_id_people:
            inquiry = f"""SELECT patients.id, departments.name, categories.name FROM patients, departments, categories
WHERE patients.id = {id_people} and patients.category = categories.id and patients.department = departments.id and patients.is_deleted = 0"""
            department_categori = self.cur.execute(inquiry).fetchone()
            if department_categori[1] not in department_by_people:
                department_by_people['удал. отделения'] += 1
            else:
                department_by_people[department_categori[1]] += 1
            if department_categori[2] not in category_by_people:
                category_by_people['удал. категории'] += 1
            else:
                category_by_people[department_categori[2]] += 1

        lines, sum_1, sum_2, records_by_category = get_detailed_report_month(self.selected_month, self.cur,
                                          main_text=f'Реабилитация за {all_month[self.selected_month[0]]}',
                                          all_doctors=all_doctors, mod=1)

        set_cell(sheet, 3, 1, 'По категориям', bold_style)
        set_cell(sheet, 3, 4, 'человек', bold_style)
        set_cell(sheet, 3, 6, 'процедур', bold_style, 2)
        set_cell(sheet, 3, 7, 'ед.', bold_style, 2)
        set_cell(sheet, 4, 1, 'Всего:', bold_style, border=border_bottom)
        set_cell(sheet, 4, 2, '', style, border=border_bottom)
        set_cell(sheet, 4, 3, '', style, border=border_r_b)
        set_cell(sheet, 4, 5, '', style, border=border_bottom)

        set_cell(sheet, 4, 4, len(right_id_people), bold_style, 2, border=border_bottom)
        set_cell(sheet, 4, 6, sum_1, bold_style, 2, border=border_bottom)
        set_cell(sheet, 4, 7, sum_2, bold_style, 2, border=border_bottom)


        if category_by_people['удал. категории'] != 0:
            all_categories.append(['удал. категории'])
        num_str = 4
        for categorie in all_categories:
            num_str += 1
            set_cell(sheet, num_str, 1, categorie[0], style)
            set_cell(sheet, num_str, 4, category_by_people[categorie[0]], style, 2, border=border_left)
            set_cell(sheet, num_str, 6, records_by_category[categorie[0]][0], style, 2)
            set_cell(sheet, num_str, 7, records_by_category[categorie[0]][1], style, 2)

        set_cell(sheet, num_str + 2, 1, 'По отделениям', bold_style)
        if department_by_people['удал. отделения'] != 0:
            all_departments.append(['удал. отделения'])
        num_str += 3
        for department in all_departments:
            num_str += 1
            set_cell(sheet, num_str, 1, department[0], style)
            set_cell(sheet, num_str, 4, department_by_people[department[0]], style)
        # ---------------------------------------------------------------------

        # ------------------------------------------------------- По процедурам

        num_str += 3
        set_cell(sheet, num_str, 1, '* данные таблице ниже о количестве человек рассчитаны другим методом (считаются все: и первичные и повторные) в отличие от верхней части отчёта (количество человек считается однократно,только первичные)и поэтому количество человек в таблице ниже при суммировании не совпадет с данными таблице выше.', pale_style)
        for line in lines:
            num_str += 1
            for i in range(1, len(line) + 1):
                call = line[i - 1]
                set_cell(sheet, num_str, i, call.text, call.style, border=call.border, alignment=call.alignment)

        num_str += 2
        date_1 = get_datetime_date(day=1, month=int(self.selected_month[0]), year=int(self.selected_month[1]))
        date_2 = date_1 + relativedelta(months=1)
        while True:
            lines = get_detailed_report_month(date_1.strftime("%d.%m.%Y"), self.cur, main_text=f'Реабилитация за {date_1.strftime("%d.%m.%Y")}', is_month=False, all_doctors=all_doctors)
            if date_1.day == date_2.day and date_1.month == date_2.month and date_1.year == date_2.year:
                break
            for line in lines:
                num_str += 1
                for i in range(1, len(line) + 1):
                    call = line[i - 1]
                    set_cell(sheet, num_str, i, call.text, call.style, border=call.border, alignment=call.alignment)
            num_str += 1
            for i in range(1, len(line) + 1):
                set_cell(sheet, num_str, i, '', style, border=border_t_b_bold, fill=color_light_gray)
            date_1 = date_1 + relativedelta(days=1)

        # ---------------------------------------------------------------------

        num_str += 3
        set_cell(sheet, num_str, 2, 'Исполнители', bold_style)
        for doctor in all_doctors:
            if doctor[3]:
                continue
            num_str += 1
            set_cell(sheet, num_str, 1, doctor[1] + '…' * 40, style)
            set_cell(sheet, num_str, 6, doctor[2], bold_style)
        num_str += 1
        set_cell(sheet, num_str, 2, 'Удалённые исполнители', bold_style)
        for doctor in all_doctors:
            if doctor[3]:
                num_str += 1
                set_cell(sheet, num_str, 1, doctor[1] + '…' * 40, style)
                set_cell(sheet, num_str, 6, doctor[2], bold_style)
        num_str += 2
        set_cell(sheet, num_str, 1, f'* Данный отчёт составлен с помощью программы "Журнал ЛФК" версии {version}', pale_style)

        sheet.column_dimensions['D'].width = 5.5
        sheet.column_dimensions['E'].width = 6

        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.orientation = 'landscape'
        sheet.set_printer_settings(9, 'landscape')

        try:
            book.save(file_name)
        except PermissionError:
            pass
        self.close()

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
