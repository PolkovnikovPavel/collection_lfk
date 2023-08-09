import os, sqlite3, openpyxl, datetime, copy
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QInputDialog, QCheckBox, QWidget, QSpinBox
from PyQt5 import QtCore, QtGui
from dateutil.relativedelta import relativedelta
from description_menu import version

from data.design.report_7 import Ui_MainWindow as Ui_Report7
from report_5 import set_cell, color_light_gray


REPORT_TITLE = 'отчёт нагрузок'


all_month = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
             '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
             '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}

bold_big_style = Font(size="14", bold=True)
bold_medium_style = Font(size="13", bold=True)
bold_style = Font(size="11", bold=True)
style = Font(size="11")
pale_style = Font(size="9", color='00555555', italic=True)


border_t_b = Border(bottom=Side(style='medium'),
                     top=Side(style='medium'))
border_left = Border(left=Side(style='medium'))
border_t_b_bold = Border(bottom=Side(style='thick'),
                     top=Side(style='thick'))
border_r_b = Border(bottom=Side(style='medium'),
                     right=Side(style='medium'))
border_bottom = Border(bottom=Side(style='medium'))
border_right = Border(right=Side(style='medium'))

class Call:
    def __init__(self, text, style=style, border=None, alignment=1):
        self.text = text
        self.style = style
        self.border = border
        self.alignment = alignment


def get_datetime_from_text(text):
    text = text.split('.')
    datetime_date = datetime.date(day=int(text[0]), month=int(text[1]), year=int(text[2]))
    return datetime_date


def get_date_calendar(calendar):
    '''
    :param calendar:
    :return: возращает выбраную дату в отформатированном виде(дд.мм.гггг)
    '''
    date = calendar.selectedDate()
    date = f'{date.day()}.{date.month()}.{date.year()}'
    if len(date.split('.')[0]) == 1:
        # проверка на то, что день записан одним символом
        date = f'0{date}'
    if len(date.split('.')[1]) == 1:
        # проверка на то, что месяц записан одним символом
        date = f'{date.split(".")[0]}.0{".".join(date.split(".")[1:])}'
    return date


def get_datetime_date(day, month, year):
    datetime_date = datetime.date(day=day, month=month,
                         year=year)
    return datetime_date


def get_num_from_date(date):
    d, m, y = date.split('.')
    return int(m) * 31 + int(y) * 385 + int(d)


def get_next_month(now):
    month, year = now
    if month < 12:
        return [month + 1, year]
    return [1, year + 1]


def get_inquiry_all_records(date):
    inquiry = f"""SELECT records.id, accounts.id, places.price, records.date FROM records, lessons, places, accounts, patients
            WHERE ({date}) and (lessons.id = records.lesson_id_1 or 
                              lessons.id = records.lesson_id_2 or
                              lessons.id = records.lesson_id_3) and 
            lessons.id_plase = places.id and lessons.id_doctor = accounts.id and records.is_deleted = 0 and patients.is_deleted = 0 and records.patient_id = patients.id"""
    return inquiry


def get_count_worked_days(period, cur):
    date_1, date_2 = period
    my1 = list(map(int, date_1.split('.')[1:]))
    my2 = list(map(int, date_2.split('.')[1:]))
    right_months = [my1]
    if my1 != my2:
        my = my1
        while my != my2:
            my = get_next_month(my)
            right_months.append(my)

    if len(right_months) == 1:
        date = f"records.date LIKE '%{right_months[0][0]}.{right_months[0][1]}'"
    else:
        date = []
        for month in right_months:
            date.append(f"records.date LIKE '%{month[0]}.{month[1]}'")
        date = ' OR '.join(date)

    inquiry = get_inquiry_all_records(date)
    all_records = cur.execute(inquiry).fetchall()
    res = {}
    all_dates = []
    date_1 = get_datetime_from_text(date_1)
    date_2 = get_datetime_from_text(date_2)

    while True:
        all_dates.append(f'{str(date_1.day).rjust(2, "0")}.{str(date_1.month).rjust(2, "0")}.{date_1.year}')
        if date_1.day == date_2.day and date_1.month == date_2.month and date_1.year == date_2.year:
            break
        date_1 = date_1 + relativedelta(days=1)

    inquiry = f"""SELECT DISTINCT id, name FROM accounts WHERE is_deleted = 0"""
    all_doctors = cur.execute(inquiry).fetchall()
    right_doctors = []

    for id, name in all_doctors:
        res[id] = set()
        right_doctors.append(id)

    for record_id, doc_id, prise, record_date in all_records:
        if doc_id in right_doctors:
            if record_date in all_dates:
                res[doc_id].add(record_date)
    for doc_id in res:
        res[doc_id] = len(res[doc_id])
    return res


def get_detailed_report_load(period, cur, right_doctors=(), tabl=None, all_doctors=()):
    date_1, date_2 = period
    my1 = list(map(int, date_1.split('.')[1:]))
    my2 = list(map(int, date_2.split('.')[1:]))
    right_months = [my1]
    if my1 != my2:
        my = my1
        while my != my2:
            my = get_next_month(my)
            right_months.append(my)

    if len(right_months) == 1:
        date = f"records.date LIKE '%{right_months[0][0]}.{right_months[0][1]}'"
    else:
        date = []
        for month in right_months:
            date.append(f"records.date LIKE '%{month[0]}.{month[1]}'")
        date = ' OR '.join(date)

    inquiry = get_inquiry_all_records(date)
    all_records = cur.execute(inquiry).fetchall()

    res = {}
    all_dates = {}
    date_1 = get_datetime_from_text(date_1)
    date_2 = get_datetime_from_text(date_2)

    while True:
        all_dates[f'{str(date_1.day).rjust(2, "0")}.{str(date_1.month).rjust(2, "0")}.{date_1.year}'] = [0, 0]
        if date_1.day == date_2.day and date_1.month == date_2.month and date_1.year == date_2.year:
            break
        date_1 = date_1 + relativedelta(days=1)

    global_data = {}
    for id in right_doctors:
        global_data[id] = [0, 0]
        res[id] = copy.deepcopy(all_dates)

    for record_id, doc_id, prise, record_date in all_records:
        if doc_id in right_doctors:
            if record_date in res[doc_id]:
                res[doc_id][record_date][0] += 1
                res[doc_id][record_date][1] += prise
                global_data[doc_id][0] += 1
                global_data[doc_id][1] += prise

    right_counts_worked_days = get_right_counts_worked_days(tabl, all_doctors)
    for id in right_doctors:
        global_data[id].append(right_counts_worked_days[id])

    return res, global_data, all_dates


def get_right_counts_worked_days(tabl, all_doctors):
    res = {}
    for row in range(len(all_doctors)):
        count = tabl.cellWidget(row, 1).value()
        res[all_doctors[row][0]] = count
    return res



class ReportMenu7(QMainWindow, Ui_Report7):
    def __init__(self, main_menu, ac_name, db_name):
        super().__init__()
        self.con = sqlite3.connect(db_name)
        self.cur = self.con.cursor()
        self.main_menu = main_menu
        self.ac_name = ac_name
        self.setupUi(self)
        self.setWindowTitle('Отчёт за месяц')
        self.choice_date_1.clicked['QDate'].connect(self.choiced_date_1)
        self.choice_date_2.clicked['QDate'].connect(self.choiced_date_2)
        self.button.clicked.connect(self.create_report)

        inquiry = f"""SELECT DISTINCT id, name, short_name FROM accounts WHERE is_deleted = 0"""
        self.all_doctors = self.cur.execute(inquiry).fetchall()
        font = QtGui.QFont()
        font.setPointSize(15)
        self.right_doctors = []

        self.scroll_doctor.setRowCount(len(self.all_doctors))
        self.scroll_doctor.setColumnCount(2)
        self.scroll_doctor.setColumnWidth(0, 700)
        self.scroll_doctor.setHorizontalHeaderLabels(['ФИО', 'раб. дней'])

        i = -1
        for doctor in self.all_doctors:
            i += 1
            cb = QCheckBox(f'{doctor[1]} ({doctor[2]})', self)
            cb.args = doctor
            cb.stateChanged.connect(self.change_doctor)
            cb.setFont(font)
            sb = QSpinBox()
            sb.setMaximum(1000)
            sb.args = doctor
            sb.setFont(font)

            self.scroll_doctor.setCellWidget(i, 0, cb)
            self.scroll_doctor.setCellWidget(i, 1, sb)

    def set_count_worked_days(self):
        date_1 = get_date_calendar(self.choice_date_1)
        date_2 = get_date_calendar(self.choice_date_2)
        data_about_worked_days = get_count_worked_days((date_1, date_2), self.cur)
        for row in range(len(self.all_doctors)):
            obj = self.scroll_doctor.cellWidget(row, 1)
            obj.setValue(data_about_worked_days[obj.args[0]])

    def choiced_date_1(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_1)) >
           get_num_from_date(get_date_calendar(self.choice_date_2))):
            self.choice_date_2.setSelectedDate(self.choice_date_1.selectedDate())
        self.set_count_worked_days()

    def choiced_date_2(self):
        if (get_num_from_date(get_date_calendar(self.choice_date_2)) <
           get_num_from_date(get_date_calendar(self.choice_date_1))):
            self.choice_date_1.setSelectedDate(self.choice_date_2.selectedDate())
        self.set_count_worked_days()

    def test_function(self, x):
        obj = self.sender()
        print(x, obj)

    def change_doctor(self, state):
        obj = self.sender()
        if state:
            self.right_doctors.append(obj.args[0])
        else:
            del self.right_doctors[self.right_doctors.index(obj.args[0])]

    def create_report(self):
        if len(self.right_doctors) == 0:
            self.error_text.setText('Выберете хотя бы одного специалиста')
            return
        file_name = QFileDialog.getSaveFileName(self, 'Путь, где сохранить отчёт', filter='*.xlsx')[0]
        if not file_name:
            return
        try:
            book = openpyxl.load_workbook(file_name)
            for sheet_name in book.sheetnames:
                if sheet_name == REPORT_TITLE:
                    sheet = book.get_sheet_by_name(sheet_name)
                    book.remove_sheet(sheet)
            sheet = book.create_sheet()
        except Exception:
            book = openpyxl.Workbook()
            sheet = book.active

        sheet.title = REPORT_TITLE

        # The data
        date_1 = get_date_calendar(self.choice_date_1)
        date_2 = get_date_calendar(self.choice_date_2)
        set_cell(sheet, 1, 1, 'Отчёт по нагрузке', bold_big_style)
        set_cell(sheet, 1, 6, f'С {date_1} по {date_2}', bold_big_style)

        doctor_data, global_doctor_data, all_dates = get_detailed_report_load((date_1, date_2), self.cur, self.right_doctors, self.scroll_doctor, self.all_doctors)

        num_str = 3
        set_cell(sheet, num_str, 1, 'за период', bold_style)
        num_str += 1
        sheet.row_dimensions[num_str].height = 32
        set_cell(sheet, num_str, 1, 'нагрузка\nпо усл.ед.', bold_style, v_alignment=2, wrap_text=True)
        set_cell(sheet, num_str, 2, 'ФИО', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 3, 'раб. дней', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 4, 'пр/день', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 5, 'ед/день', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 6, 'норма/день', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 7, 'пр/период', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 8, 'ед/период', bold_style, v_alignment=2)
        set_cell(sheet, num_str, 9, 'норма/период', bold_style, v_alignment=2)

        inquiry = f"""SELECT DISTINCT id, name, load_rate FROM accounts WHERE is_deleted = 0"""
        all_doctors = self.cur.execute(inquiry).fetchall()
        for id, name, load_rate in all_doctors:
            if id in global_doctor_data:
                num_str += 1
                if global_doctor_data[id][2] != 0:
                    set_cell(sheet, num_str, 1, f'{round((global_doctor_data[id][1] / (load_rate * global_doctor_data[id][2])) * 100)}%', style)
                else:
                    set_cell(sheet, num_str, 1, '0%', style)
                set_cell(sheet, num_str, 2, name, style)
                set_cell(sheet, num_str, 3, str(global_doctor_data[id][2]), style, 2)
                if global_doctor_data[id][2] != 0:
                    set_cell(sheet, num_str, 4, str(round(global_doctor_data[id][0] / global_doctor_data[id][2], 1)), style, 2)
                    set_cell(sheet, num_str, 5, str(round(global_doctor_data[id][1] / global_doctor_data[id][2], 1)), style, 2)
                else:
                    set_cell(sheet, num_str, 4, '0.0', style, 2)
                    set_cell(sheet, num_str, 5, '0.0', style, 2)
                set_cell(sheet, num_str, 6, str(load_rate), style, 2)
                set_cell(sheet, num_str, 7, str(global_doctor_data[id][0]), style, 2)
                set_cell(sheet, num_str, 8, str(global_doctor_data[id][1]), style, 2)
                set_cell(sheet, num_str, 9, str(load_rate * global_doctor_data[id][2]), style, 2)
        num_str += 3
        for date in all_dates:
            set_cell(sheet, num_str, 1, date, bold_style)
            num_str += 1
            set_cell(sheet, num_str, 1, 'нагрузка', bold_style)
            set_cell(sheet, num_str, 2, 'ФИО', bold_style)
            set_cell(sheet, num_str, 3, 'процедур', bold_style)
            set_cell(sheet, num_str, 4, 'усл. ед.', bold_style)
            set_cell(sheet, num_str, 5, 'норма', bold_style)

            for id, name, load_rate in all_doctors:
                if id in doctor_data:
                    num_str += 1
                    set_cell(sheet, num_str, 1, f'{round((doctor_data[id][date][1] / load_rate) * 100)}%', style)
                    set_cell(sheet, num_str, 2, name, style)
                    set_cell(sheet, num_str, 3, str(doctor_data[id][date][0]), style, 2)
                    set_cell(sheet, num_str, 4, str(doctor_data[id][date][1]), style, 2)
                    set_cell(sheet, num_str, 5, str(load_rate), style, 2)
            num_str += 2

        num_str += 2
        set_cell(sheet, num_str, 1, f'* Данный отчёт составлен с помощью программы "Журнал ЛФК" версии {version}', pale_style)

        sheet.column_dimensions['A'].width = 11
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 8
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 10.2
        sheet.column_dimensions['H'].width = 10.2
        sheet.column_dimensions['I'].width = 14.5




        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.set_printer_settings(9, 'landscape')
        sheet.orientation = 'landscape'

        try:
            book.save(file_name)
        except PermissionError:
            pass
        self.close()

    def open_main_menu(self):
        self.close()  # закрывает это окно
        self.main_menu.show()
